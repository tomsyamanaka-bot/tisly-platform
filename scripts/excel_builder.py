"""TOMS 見積生成 — Excel テンプレート読込・転記"""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from config import TomsConfig
from notion_client import EstimateData

# TOMS 標準見積フォーマット セルマッピング（1-based）
CELL_COMPANY = (1, 1)          # A1  会社名
CELL_ISSUE_DATE = (1, 7)       # G1  発行日
CELL_ESTIMATE_NO = (2, 7)      # G2  見積番号
CELL_TITLE = (3, 1)            # A3  見 積 書
CELL_CUSTOMER = (6, 3)         # C6  宛名
CELL_PERSON = (7, 1)           # A7  担当
CELL_PROJECT_LABEL = (8, 1)    # A8  件名ラベル
CELL_PROJECT = (9, 4)          # D9  件名
CELL_HEADER_TOTAL_LABEL = (16, 1)
CELL_HEADER_TOTAL = (17, 4)    # D17 税込合計（ヘッダー部）

ITEM_HEADER_ROW = 11
ITEM_START_ROW = 12
ITEM_COL_NO = 1
ITEM_COL_NAME = 2
ITEM_COL_QTY = 3
ITEM_COL_UNIT_PRICE = 4
ITEM_COL_AMOUNT = 5

SUMMARY_SUBTOTAL_ROW = 47
SUMMARY_TAX_ROW = 48
SUMMARY_TOTAL_ROW = 49
SUMMARY_LABEL_COL = 6
SUMMARY_VALUE_COL = 7

REMARKS_START_ROW = 51


def ensure_template(path: Path, company_name: str) -> Path:
    """テンプレートが無ければ TOMS 標準レイアウトで作成する。"""
    if path.is_file():
        return path

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "見積書"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14

    ws.cell(*CELL_COMPANY, company_name).font = Font(bold=True, size=12)
    ws.cell(*CELL_TITLE, "見 積 書").font = Font(bold=True, size=16)
    ws.cell(*CELL_TITLE).alignment = center

    ws.cell(1, 6, "発行日")
    ws.cell(2, 6, "見積番号")
    ws.cell(*CELL_PROJECT_LABEL, "件名")
    ws.cell(*CELL_HEADER_TOTAL_LABEL, "税込合計")

    for col, label in zip(
        range(ITEM_COL_NO, ITEM_COL_AMOUNT + 1),
        ("No", "項目", "数量", "単価", "金額"),
    ):
        cell = ws.cell(ITEM_HEADER_ROW, col, label)
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = border

    for row, label in zip(
        (SUMMARY_SUBTOTAL_ROW, SUMMARY_TAX_ROW, SUMMARY_TOTAL_ROW),
        ("小計", "消費税", "税込合計"),
    ):
        ws.cell(row, SUMMARY_LABEL_COL, label).font = Font(bold=True)
        ws.cell(row, SUMMARY_VALUE_COL).alignment = right
        ws.cell(row, SUMMARY_VALUE_COL).number_format = "#,##0"

    ws.cell(REMARKS_START_ROW, 1, "〈備考〉").font = Font(bold=True)

    wb.save(path)
    return path


def _format_yen(value: int) -> str:
    return f"¥{value:,}"


def build_estimate_xlsx(
    data: EstimateData,
    config: TomsConfig,
    output_path: Path,
) -> Path:
    """Notion データを TOMS 標準見積フォーマットへ転記して保存する。"""
    template_path = ensure_template(config.template_path, config.company_name)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb["見積書"]

    ws.cell(*CELL_COMPANY, config.company_name)
    ws.cell(*CELL_ISSUE_DATE, data.issue_date)
    ws.cell(*CELL_ESTIMATE_NO, data.estimate_no)
    ws.cell(*CELL_CUSTOMER, f"{data.customer_name} 御中")
    ws.cell(*CELL_PERSON, f"担当: {data.person_in_charge}")
    ws.cell(*CELL_PROJECT, data.case_name)
    ws.cell(*CELL_HEADER_TOTAL, _format_yen(data.total))

    for idx, item in enumerate(data.items):
        row = ITEM_START_ROW + idx
        ws.cell(row, ITEM_COL_NO, idx + 1)
        ws.cell(row, ITEM_COL_NAME, item.item_name)
        ws.cell(row, ITEM_COL_QTY, item.quantity)
        ws.cell(row, ITEM_COL_UNIT_PRICE, item.unit_price).number_format = "#,##0"
        ws.cell(row, ITEM_COL_AMOUNT, item.amount).number_format = "#,##0"

    ws.cell(SUMMARY_SUBTOTAL_ROW, SUMMARY_VALUE_COL, data.subtotal).number_format = "#,##0"
    ws.cell(SUMMARY_TAX_ROW, SUMMARY_VALUE_COL, data.tax).number_format = "#,##0"
    ws.cell(SUMMARY_TOTAL_ROW, SUMMARY_VALUE_COL, data.total).number_format = "#,##0"

    ws.cell(REMARKS_START_ROW, 1, "〈備考〉")
    default_remarks = data.remarks or [
        "・価格は税抜単価に基づき算出しています",
        "・正式見積前に現地確認が必要な場合があります",
    ]
    for i, remark in enumerate(default_remarks):
        ws.cell(REMARKS_START_ROW + 1 + i, 1, remark)

    wb.save(output_path)
    return output_path


def resolve_output_dir(data: EstimateData, config: TomsConfig) -> Path:
    """output/{顧客名}/{案件名}/{YYYY-MM}/ を返す。"""
    ym = data.issue_date[:7] if len(data.issue_date) >= 7 else data.issue_date
    return config.output_dir / data.customer_folder / data.case_name / ym
