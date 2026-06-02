"""TOMS Phase 2-2 — 請求書 Excel テンプレート読込・転記"""

from __future__ import annotations

import shutil
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from config import TomsConfig
from notion_client import InvoiceData

CELL_COMPANY = (1, 1)
CELL_ISSUE_DATE = (1, 7)
CELL_INVOICE_NO = (2, 7)
CELL_TITLE = (3, 1)
CELL_CUSTOMER = (6, 3)
CELL_PERSON = (7, 1)
CELL_PROJECT = (9, 4)
CELL_DUE_DATE = (10, 4)
CELL_BANK = (11, 1)

ITEM_HEADER_ROW = 13
ITEM_START_ROW = 14
ITEM_COL_NO = 1
ITEM_COL_NAME = 2
ITEM_COL_QTY = 3
ITEM_COL_UNIT_PRICE = 4
ITEM_COL_AMOUNT = 5

SUMMARY_SUBTOTAL_ROW = 47
SUMMARY_TAX_ROW = 48
SUMMARY_TOTAL_ROW = 49
SUMMARY_LABEL_COL = 6
SUMMARY_VALUE_COL = 7

REMARKS_START_ROW = 51


def default_due_date(issue_date: str) -> str:
    """請求日から翌月末を支払期限とする。"""
    base = date.fromisoformat(issue_date)
    if base.month == 12:
        end = date(base.year + 1, 1, 1)
    else:
        end = date(base.year, base.month + 1, 1)
    last = end - timedelta(days=1)
    return last.isoformat()


def ensure_invoice_template(path: Path, company_name: str) -> Path:
    if path.is_file():
        return path

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "請求書"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14

    ws.cell(*CELL_COMPANY, company_name).font = Font(bold=True, size=12)
    ws.cell(*CELL_TITLE, "請 求 書").font = Font(bold=True, size=16)
    ws.cell(*CELL_TITLE).alignment = center

    ws.cell(1, 6, "請求日")
    ws.cell(2, 6, "請求番号")
    ws.cell(8, 1, "件名")
    ws.cell(9, 1, "支払期限")

    for col, label in zip(
        range(ITEM_COL_NO, ITEM_COL_AMOUNT + 1),
        ("No", "項目", "数量", "単価", "金額"),
    ):
        cell = ws.cell(ITEM_HEADER_ROW, col, label)
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = border

    for row, label in zip(
        (SUMMARY_SUBTOTAL_ROW, SUMMARY_TAX_ROW, SUMMARY_TOTAL_ROW),
        ("小計", "消費税", "税込合計"),
    ):
        ws.cell(row, SUMMARY_LABEL_COL, label).font = Font(bold=True)
        ws.cell(row, SUMMARY_VALUE_COL).alignment = right
        ws.cell(row, SUMMARY_VALUE_COL).number_format = "#,##0"

    ws.cell(REMARKS_START_ROW, 1, "〈備考〉").font = Font(bold=True)
    wb.save(path)
    return path


def _format_yen(value: int) -> str:
    return f"¥{value:,}"


def build_invoice_xlsx(
    data: InvoiceData,
    config: TomsConfig,
    output_path: Path,
) -> Path:
    template_path = ensure_invoice_template(config.invoice_template_path, config.company_name)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb["請求書"]

    due = data.due_date or default_due_date(data.issue_date)

    ws.cell(*CELL_COMPANY, config.company_name)
    ws.cell(*CELL_ISSUE_DATE, data.issue_date)
    ws.cell(*CELL_INVOICE_NO, data.invoice_no)
    ws.cell(*CELL_CUSTOMER, f"{data.customer_name} 御中")
    ws.cell(*CELL_PERSON, f"担当: {data.person_in_charge}")
    ws.cell(*CELL_PROJECT, data.case_name)
    ws.cell(*CELL_DUE_DATE, due)
    ws.cell(*CELL_BANK, f"振込先: {config.bank_info}")

    for idx, item in enumerate(data.items):
        row = ITEM_START_ROW + idx
        ws.cell(row, ITEM_COL_NO, idx + 1)
        ws.cell(row, ITEM_COL_NAME, item.item_name)
        ws.cell(row, ITEM_COL_QTY, item.quantity)
        ws.cell(row, ITEM_COL_UNIT_PRICE, item.unit_price).number_format = "#,##0"
        ws.cell(row, ITEM_COL_AMOUNT, item.amount).number_format = "#,##0"

    ws.cell(SUMMARY_SUBTOTAL_ROW, SUMMARY_VALUE_COL, data.subtotal).number_format = "#,##0"
    ws.cell(SUMMARY_TAX_ROW, SUMMARY_VALUE_COL, data.tax).number_format = "#,##0"
    ws.cell(SUMMARY_TOTAL_ROW, SUMMARY_VALUE_COL, data.total).number_format = "#,##0"

    ws.cell(REMARKS_START_ROW, 1, "〈備考〉")
    remarks = data.remarks or [
        "・上記の通りご請求申し上げます",
        f"・お振込手数料は貴社にてご負担ください",
        f"・案件番号: {data.case_number}",
    ]
    for i, remark in enumerate(remarks):
        ws.cell(REMARKS_START_ROW + 1 + i, 1, remark)

    wb.save(output_path)
    return output_path


def resolve_invoice_output_dir(data: InvoiceData, config: TomsConfig) -> Path:
    ym = data.issue_date[:7] if len(data.issue_date) >= 7 else data.issue_date
    return config.output_dir / data.customer_folder / data.case_name / ym
